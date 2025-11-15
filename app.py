# app.py — complete updated version
import os
import io
import json
import re
import zipfile
from datetime import datetime, timedelta
from flask import (
    Flask, request, render_template, jsonify, send_from_directory, send_file,
    redirect, url_for, session, flash
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from dotenv import load_dotenv
import smtplib
from email.message import EmailMessage

# --------- LOAD ENV ---------
load_dotenv()

SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "warulesuraj12@gmail.com")  # sender
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")  # your 16-char app password
SECRET_KEY = os.getenv("SECRET_KEY", "change_this_secret_in_prod")

# --------- CONFIG ---------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXT = {'.pdf', '.doc', '.docx'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(BASE_DIR, 'interview.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# --------- MODELS ---------
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200))
    email = db.Column(db.String(200), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(32), nullable=False)  # HR, HOD, UnitHead
    is_active = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)

class Application(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    address = db.Column(db.String(500), nullable=False)
    contact = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(200))
    # JSON columns for complex tables
    academic_json = db.Column(db.Text)
    professional_json = db.Column(db.Text)
    family_json = db.Column(db.Text)
    position_applied = db.Column(db.String(200))
    # OPTIONAL department field — useful for department-wise stats
    department = db.Column(db.String(200), default="")
    area_of_interest = db.Column(db.String(200))
    current_salary = db.Column(db.String(100))
    expected_salary = db.Column(db.String(100))
    notice_period = db.Column(db.String(100))
    resume_filename = db.Column(db.String(400))
    other_details = db.Column(db.Text)
    reference_type = db.Column(db.String(50))
    reference_name = db.Column(db.String(200))
    status = db.Column(db.String(50), default='Applied')  # Applied, Assigned, Interviewed, OnHold, Rejected, Selected, Offered, Joined
    assigned_hod_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    applied_at = db.Column(db.DateTime, default=datetime.utcnow)

# --------- HELPERS ---------
def allowed_file(filename):
    ext = os.path.splitext(filename)[1].lower()
    return ext in ALLOWED_EXT

def login_required(role=None):
    def wrapper(fn):
        def wrapped(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if role and session.get('role') != role:
                flash("Access denied", "danger")
                return redirect(url_for('login'))
            return fn(*args, **kwargs)
        wrapped.__name__ = fn.__name__
        return wrapped
    return wrapper

# --------- EMAIL UTILITIES ---------
def send_email(to, subject, html_body, cc=None):
    """
    Send HTML email using SMTP credentials from env.
    - to: string or list
    - cc: optional string or list
    Returns True on success, False on failure (and logs errors).
    """
    try:
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = SMTP_USERNAME

        if isinstance(to, (list, tuple)):
            msg['To'] = ', '.join(to)
        else:
            msg['To'] = to

        if cc:
            if isinstance(cc, (list, tuple)):
                msg['Cc'] = ', '.join(cc)
            else:
                msg['Cc'] = cc

        # plain fallback and html alternative
        msg.set_content("This email contains an HTML version. Please view in an HTML-capable client.")
        msg.add_alternative(html_body, subtype='html')

        # connect and send
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(msg)

        app.logger.info(f"Email sent to {to} subject={subject}")
        return True

    except smtplib.SMTPAuthenticationError as e:
        app.logger.error("SMTP Authentication error: %s", e, exc_info=e)
        return False
    except smtplib.SMTPResponseException as e:
        # This will include the (334, b'...') response you saw.
        app.logger.error("SMTP Response Exception: %s %s", e.smtp_code, e.smtp_error, exc_info=e)
        return False
    except Exception as e:
        app.logger.error("SMTP general exception: %s", e, exc_info=e)
        return False

def render_and_send(template_name, to_email, subject, cc=None, **template_ctx):
    """
    Render a template inside templates/email_templates/<template_name>
    and send it.
    """
    try:
        html = render_template(f"email_templates/{template_name}", **template_ctx)
    except Exception as e:
        app.logger.error("Template render failed: %s", e, exc_info=e)
        html = "<p>Notification from Accord Interview System</p>"
    return send_email(to_email, subject, html, cc=cc)

# --------- STARTUP (create DB + default HR) ---------
@app.before_first_request
def create_tables():
    db.create_all()
    hr = User.query.filter_by(email='hr@accordhospitals.com').first()
    if not hr:
        hr = User(name='HR Admin', email='hr@accordhospitals.com', role='HR', is_active=True)
        hr.set_password('ChangeMe123!')
        db.session.add(hr)
        db.session.commit()

# --------- ROUTES: homepage/auth ---------
@app.route('/')
def home():
    return render_template('home.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        password = request.form.get('password','').strip()
        user = User.query.filter_by(email=email).first()
        if not user or not user.check_password(password):
            flash("Invalid credentials", "danger")
            return render_template('login.html')
        if user.role == 'HOD' and not user.is_active:
            flash("HOD account pending HR approval", "warning")
            return render_template('login.html')
        session['user_id'] = user.id
        session['email'] = user.email
        session['role'] = user.role
        session['name'] = user.name
        if user.role == 'HR':
            return redirect(url_for('hr_dashboard'))
        if user.role == 'HOD':
            return redirect(url_for('hod_dashboard'))
        if user.role == 'UnitHead':
            return redirect(url_for('unit_dashboard'))
        return redirect(url_for('home'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("Logged out", "info")
    return redirect(url_for('login'))

@app.route('/register', methods=['GET','POST'])
def register():
    # HOD self-register
    if request.method == 'POST':
        name = request.form.get('name','').strip()
        email = request.form.get('email','').strip().lower()
        password = request.form.get('password','').strip()
        if not email or not password:
            flash("Email and password required", "danger")
            return render_template('register.html')
        if User.query.filter_by(email=email).first():
            flash("Email already registered", "warning")
            return render_template('register.html')
        user = User(name=name, email=email, role='HOD', is_active=False)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash("Registered. Await HR approval.", "success")
        return redirect(url_for('login'))
    return render_template('register.html')

# --------- APPLY (candidate) ---------
@app.route('/apply', methods=['GET','POST'])
def apply():
    if request.method == 'POST':
        name = request.form.get('name','').strip()
        address = request.form.get('address','').strip()
        contact = re.sub(r'\D', '', request.form.get('contact','').strip())
        email = request.form.get('email','').strip().lower() or None
        department = request.form.get('department','').strip() or ""

        if not name or not address or not contact or len(contact) < 10:
            flash("Name, address and valid 10-digit contact are required.", "danger")
            return render_template('candidate_form.html')

        if email and Application.query.filter_by(email=email).first():
            flash("You already applied with this email.", "warning")
            return render_template('candidate_form.html')

        if Application.query.filter_by(contact=contact).first():
            flash("You already applied with this contact number.", "warning")
            return render_template('candidate_form.html')

        academic = request.form.get('academic_json','[]')
        professional = request.form.get('professional_json','[]')
        family = request.form.get('family_json','[]')

        resume = request.files.get('resume')
        resume_filename = None
        if resume:
            fname = secure_filename(resume.filename)
            if not allowed_file(fname):
                flash("Resume must be pdf/doc/docx", "danger")
                return render_template('candidate_form.html')
            content = resume.read()
            if len(content) > MAX_FILE_SIZE:
                flash("Resume exceeds 5MB", "danger")
                return render_template('candidate_form.html')
            save_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{fname}"
            with open(os.path.join(UPLOAD_DIR, save_name),'wb') as fh:
                fh.write(content)
            resume_filename = save_name

        new_app = Application(
            name=name, address=address, contact=contact, email=email,
            academic_json=academic, professional_json=professional, family_json=family,
            position_applied=request.form.get('positionApplied'),
            department=department,
            area_of_interest=request.form.get('areaOfInterest'),
            current_salary=request.form.get('currentSalary'),
            expected_salary=request.form.get('expectedSalary'),
            notice_period=request.form.get('noticePeriod'),
            resume_filename=resume_filename,
            other_details=request.form.get('otherDetails'),
            reference_type=request.form.get('referenceType'),
            reference_name=request.form.get('referenceName')
        )
        db.session.add(new_app)
        db.session.commit()

        # Send emails: candidate + HR
        hr_user = User.query.filter_by(role='HR').first()
        hr_email = hr_user.email if hr_user else None
        if email:
            render_and_send("candidate_applied.html", email, "Accord Hospitals — Application Received", name=name, position=new_app.position_applied)
        if hr_email:
            render_and_send("candidate_applied.html", hr_email, "New Application Received — Accord Hospitals", name=name, position=new_app.position_applied)

        flash("Application submitted.", "success")
        return render_template('submitted.html', name=name)

    # GET show form
    return render_template('candidate_form.html')

# --------- HR dashboard + endpoints for charts ---------
@app.route('/hr')
@login_required(role='HR')
def hr_dashboard():
    # full candidate list for table
    apps = Application.query.order_by(Application.applied_at.desc()).all()
    # active hods
    active_hods = User.query.filter_by(role='HOD', is_active=True).all()
    # summary counts
    total = Application.query.count()
    interviewed = Application.query.filter_by(status='Interviewed').count()
    selected = Application.query.filter_by(status='Selected').count()
    rejected = Application.query.filter_by(status='Rejected').count()
    offered = Application.query.filter_by(status='Offered').count()
    onhold = Application.query.filter_by(status='OnHold').count()
    applied = Application.query.filter_by(status='Applied').count()

    return render_template('hr_dashboard.html',
                           candidates=apps,
                           active_hods=active_hods,
                           total=total, applied=applied, interviewed=interviewed,
                           selected=selected, rejected=rejected, offered=offered, onhold=onhold)

@app.route('/api/hr/department_stats')
@login_required(role='HR')
def api_hr_department_stats():
    """
    Returns JSON with department-wise counts:
    {
      "departments": ["Cardiology","Radiology"],
      "applied": [10,5],
      "interviewed": [...],
      "selected": [...]
    }
    """
    # gather list of departments
    rows = db.session.query(Application.department).all()
    departments = sorted({(r[0] or "Unspecified") for r in rows})
    data = {"departments": [], "applied": [], "interviewed": [], "selected": [], "rejected": [], "onhold": [], "offered": []}
    for d in departments:
        data["departments"].append(d)
        data["applied"].append(Application.query.filter_by(department=d, status='Applied').count())
        data["interviewed"].append(Application.query.filter_by(department=d, status='Interviewed').count())
        data["selected"].append(Application.query.filter_by(department=d, status='Selected').count())
        data["rejected"].append(Application.query.filter_by(department=d, status='Rejected').count())
        data["onhold"].append(Application.query.filter_by(department=d, status='OnHold').count())
        data["offered"].append(Application.query.filter_by(department=d, status='Offered').count())
    return jsonify(data)

# --------- HR endpoints for approve/assign/reset etc. ---------
@app.route('/hr/approve_hod/<int:hod_id>', methods=['POST'])
@login_required(role='HR')
def hr_approve_hod(hod_id):
    hod = User.query.get(hod_id)
    if not hod or hod.role != 'HOD':
        flash("HOD not found", "danger")
        return redirect(url_for('hr_dashboard'))
    hod.is_active = True
    db.session.commit()
    flash("HOD approved", "success")
    return redirect(url_for('hr_dashboard'))

@app.route('/hr/reset_password', methods=['POST'])
@login_required(role='HR')
def hr_reset_password():
    email = request.form.get('email','').strip().lower()
    new_pw = request.form.get('new_password','').strip()
    if not email or not new_pw:
        flash("Provide email and new password", "danger")
        return redirect(url_for('hr_dashboard'))
    hod = User.query.filter_by(email=email, role='HOD').first()
    if not hod:
        flash("HOD not found", "danger")
        return redirect(url_for('hr_dashboard'))
    hod.set_password(new_pw)
    db.session.commit()
    render_and_send("status_update.html", hod.email, "Your HOD account password has been reset", name=hod.name, status=f"Your password was reset. New password: {new_pw}")
    flash("Password reset and email sent", "success")
    return redirect(url_for('hr_dashboard'))

@app.route('/hr/assign/<int:app_id>', methods=['POST'])
@login_required(role='HR')
def hr_assign(app_id):
    hod_id = request.form.get('hod_id')
    app_row = Application.query.get(app_id)
    if not app_row:
        flash("Application not found", "danger")
        return redirect(url_for('hr_dashboard'))
    if not hod_id:
        flash("Select HOD", "warning")
        return redirect(url_for('hr_dashboard'))
    app_row.assigned_hod_id = int(hod_id)
    app_row.status = 'Assigned'
    db.session.commit()
    hod = User.query.get(int(hod_id))
    hr_user = User.query.filter_by(role='HR').first()
    hr_email = hr_user.email if hr_user else None
    if hod:
        render_and_send("hod_assignment.html", hod.email, "New Candidate Assigned", cc=hr_email, hod_name=hod.name, candidate=app_row)
    flash("Assigned and HOD notified", "success")
    return redirect(url_for('hr_dashboard'))

# --------- HOD dashboard & actions ---------
@app.route('/hod')
@login_required(role='HOD')
def hod_dashboard():
    # --------- UNIT HEAD DASHBOARD (WITH CHARTS & STATS) ---------

    user_id = session.get('user_id')
    apps = Application.query.filter_by(assigned_hod_id=user_id).order_by(Application.applied_at.desc()).all()
    return render_template('hod_dashboard.html', candidates=apps)

@app.route('/hod/result/<int:app_id>', methods=['POST'])
@login_required(role='HOD')
def hod_result(app_id):
    status = request.form.get('status')
    if status not in ('Interviewed', 'Rejected', 'Selected', 'OnHold'):
        flash("Invalid status", "danger")
        return redirect(url_for('hod_dashboard'))
    app_row = Application.query.get(app_id)
    if not app_row or app_row.assigned_hod_id != session.get('user_id'):
        flash("Not assigned to you", "danger")
        return redirect(url_for('hod_dashboard'))
    app_row.status = status
    db.session.commit()
    # notify candidate + cc HR
    hr_user = User.query.filter_by(role='HR').first()
    hr_email = hr_user.email if hr_user else None
    if app_row.email:
        render_and_send("status_update.html", app_row.email, f"Application Status: {app_row.status}", cc=hr_email, name=app_row.name, status=app_row.status)
    flash("Status updated and candidate notified", "success")
    return redirect(url_for('hod_dashboard'))
# --------- UNIT HEAD DASHBOARD (WITH CHARTS & STATS) ---------
@app.route('/unit')
@login_required(role='UnitHead')
def unit_dashboard():
    # Total stats
    total = Application.query.count()
    applied = Application.query.filter_by(status='Applied').count()
    interviewed = Application.query.filter_by(status='Interviewed').count()
    selected = Application.query.filter_by(status='Selected').count()
    rejected = Application.query.filter_by(status='Rejected').count()
    onhold = Application.query.filter_by(status='OnHold').count()
    joined = Application.query.filter_by(status='Joined').count()

    # --------- DEPARTMENT-WISE STATS ---------
    dept_rows = db.session.query(
        Application.department,
        db.func.count(Application.id)
    ).group_by(Application.department).all()

    dept_labels = [r[0] if r[0] else "Unspecified" for r in dept_rows]
    dept_values = [r[1] for r in dept_rows]

    return render_template(
        'unit_dashboard.html',
        total=total,
        applied=applied,
        interviewed=interviewed,
        selected=selected,
        rejected=rejected,
        onhold=onhold,
        joined=joined,
        dept_labels=json.dumps(dept_labels),
        dept_values=json.dumps(dept_values)
    )

# --------- Candidate view ---------
@app.route('/candidate/<int:app_id>')
@login_required()
def candidate_view(app_id):
    a = Application.query.get(app_id)
    if not a:
        flash("Candidate not found", "danger")
        return redirect(url_for('home'))
    try: academic = json.loads(a.academic_json or '[]')
    except: academic = []
    try: professional = json.loads(a.professional_json or '[]')
    except: professional = []
    try: family = json.loads(a.family_json or '[]')
    except: family = []
    return render_template('candidate_view.html', a=a, academic=academic, professional=professional, family=family)

# --------- EXPORTS: Excel / ZIP / Filtered ---------
@app.route('/hr/export_excel')
@login_required(role='HR')
def export_excel():
    apps = Application.query.order_by(Application.applied_at.asc()).all()
    wb = Workbook()
    ws_apps = wb.active
    ws_apps.title = "Applications"
    ws_apps.append(["Application ID","Name","Address","Contact","Email","Position Applied","Department","Area of Interest","Current Salary","Expected Salary","Notice Period","Status","Assigned HOD","Resume Filename","Reference Type","Reference Name","Other Details","Applied At"])
    ws_acad = wb.create_sheet("Academic")
    ws_acad.append(["Application ID","Sr.No","Qualification","College/University","Year of Passing","Grade/Percentage"])
    ws_prof = wb.create_sheet("Professional")
    ws_prof.append(["Application ID","Sr.No","Name of Company","Designation","Work Tenure","Reason for leaving"])
    ws_fam = wb.create_sheet("Family")
    ws_fam.append(["Application ID","Sr.No","Name","Relation","Age","Occupation"])
    for a in apps:
        hod = User.query.filter_by(id=a.assigned_hod_id).first()
        hod_name = hod.name if hod else ""
        ws_apps.append([a.id, a.name, a.address, a.contact, a.email, a.position_applied, a.department, a.area_of_interest, a.current_salary, a.expected_salary, a.notice_period, a.status, hod_name, a.resume_filename, a.reference_type, a.reference_name, (a.other_details or ""), a.applied_at.strftime("%Y-%m-%d %H:%M")])
        try: acad_list = json.loads(a.academic_json or "[]")
        except: acad_list = []
        if isinstance(acad_list, dict): acad_list = [acad_list]
        for r in acad_list:
            ws_acad.append([a.id, r.get("sr",""), r.get("qualification",""), r.get("college",""), r.get("year",""), r.get("grade","")])
        try: prof_list = json.loads(a.professional_json or "[]")
        except: prof_list = []
        if isinstance(prof_list, dict): prof_list = [prof_list]
        for r in prof_list:
            ws_prof.append([a.id, r.get("sr",""), r.get("company",""), r.get("designation",""), r.get("tenure",""), r.get("reason","")])
        try: fam_list = json.loads(a.family_json or "[]")
        except: fam_list = []
        if isinstance(fam_list, dict): fam_list = [fam_list]
        for r in fam_list:
            ws_fam.append([a.id, r.get("sr",""), r.get("name",""), r.get("relation",""), r.get("age",""), r.get("occupation","")])
    out_path = os.path.join(BASE_DIR, "applications_export_structured.xlsx")
    wb.save(out_path)
    return send_file(out_path, as_attachment=True)

@app.route('/hr/export_zip')
@login_required(role='HR')
def export_zip():
    apps = Application.query.order_by(Application.applied_at.asc()).all()
    wb = Workbook()
    ws_apps = wb.active
    ws_apps.title = "Applications"
    ws_apps.append(["Application ID","Name","Address","Contact","Email","Position Applied","Department","Area of Interest","Current Salary","Expected Salary","Notice Period","Status","Assigned HOD","Resume Filename","Reference Type","Reference Name","Other Details","Applied At"])
    ws_acad = wb.create_sheet("Academic")
    ws_acad.append(["Application ID","Sr.No","Qualification","College/University","Year","Grade"])
    ws_prof = wb.create_sheet("Professional")
    ws_prof.append(["Application ID","Sr.No","Company","Designation","Tenure","Reason"])
    ws_fam = wb.create_sheet("Family")
    ws_fam.append(["Application ID","Sr.No","Name","Relation","Age","Occupation"])
    for a in apps:
        hod = User.query.filter_by(id=a.assigned_hod_id).first()
        hod_name = hod.name if hod else ""
        ws_apps.append([a.id, a.name, a.address, a.contact, a.email, a.position_applied, a.department, a.area_of_interest, a.current_salary, a.expected_salary, a.notice_period, a.status, hod_name, a.resume_filename, a.reference_type, a.reference_name, (a.other_details or ""), a.applied_at.strftime("%Y-%m-%d %H:%M")])
        try: acad_list = json.loads(a.academic_json or "[]")
        except: acad_list = []
        if isinstance(acad_list, dict): acad_list = [acad_list]
        for r in acad_list:
            ws_acad.append([a.id, r.get("sr",""), r.get("qualification",""), r.get("college",""), r.get("year",""), r.get("grade","")])
        try: prof_list = json.loads(a.professional_json or "[]")
        except: prof_list = []
        if isinstance(prof_list, dict): prof_list = [prof_list]
        for r in prof_list:
            ws_prof.append([a.id, r.get("sr",""), r.get("company",""), r.get("designation",""), r.get("tenure",""), r.get("reason","")])
        try: fam_list = json.loads(a.family_json or "[]")
        except: fam_list = []
        if isinstance(fam_list, dict): fam_list = [fam_list]
        for r in fam_list:
            ws_fam.append([a.id, r.get("sr",""), r.get("name",""), r.get("relation",""), r.get("age",""), r.get("occupation","")])
    excel_mem = io.BytesIO()
    wb.save(excel_mem)
    excel_mem.seek(0)
    zip_mem = io.BytesIO()
    with zipfile.ZipFile(zip_mem, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("applications_export_structured.xlsx", excel_mem.read())
        for a in apps:
            if a.resume_filename:
                path = os.path.join(UPLOAD_DIR, a.resume_filename)
                if os.path.exists(path):
                    zf.write(path, arcname=f"resumes/{a.resume_filename}")
    zip_mem.seek(0)
    return send_file(zip_mem, mimetype="application/zip", as_attachment=True, download_name="export_bundle.zip")

@app.route('/hr/export_filtered')
@login_required(role='HR')
def export_filtered():
    from_date_str = request.args.get('from_date', '').strip()
    to_date_str = request.args.get('to_date', '').strip()
    status = request.args.get('status', '').strip()
    if not from_date_str or not to_date_str:
        flash("Please select From and To dates.", "warning")
        return redirect(url_for('hr_dashboard'))
    try:
        start_dt = datetime.strptime(from_date_str, "%Y-%m-%d")
        end_dt = datetime.strptime(to_date_str, "%Y-%m-%d") + timedelta(days=1)
    except Exception:
        flash("Invalid date format.", "danger")
        return redirect(url_for('hr_dashboard'))
    q = Application.query.filter(Application.applied_at >= start_dt, Application.applied_at < end_dt)
    if status:
        q = q.filter_by(status=status)
    apps = q.order_by(Application.applied_at.asc()).all()
    wb = Workbook()
    ws_apps = wb.active
    ws_apps.title = "Applications"
    ws_apps.append(["Application ID","Name","Address","Contact","Email","Position Applied","Department","Area of Interest","Current Salary","Expected Salary","Notice Period","Status","Assigned HOD","Resume Filename","Reference Type","Reference Name","Other Details","Applied At"])
    ws_acad = wb.create_sheet("Academic")
    ws_acad.append(["Application ID","Sr.No","Qualification","College/University","Year","Grade"])
    ws_prof = wb.create_sheet("Professional")
    ws_prof.append(["Application ID","Sr.No","Company","Designation","Tenure","Reason"])
    ws_fam = wb.create_sheet("Family")
    ws_fam.append(["Application ID","Sr.No","Name","Relation","Age","Occupation"])
    for a in apps:
        hod = User.query.filter_by(id=a.assigned_hod_id).first()
        hod_name = hod.name if hod else ""
        ws_apps.append([a.id, a.name, a.address, a.contact, a.email, a.position_applied, a.department, a.area_of_interest, a.current_salary, a.expected_salary, a.notice_period, a.status, hod_name, a.resume_filename, a.reference_type, a.reference_name, (a.other_details or ""), a.applied_at.strftime("%Y-%m-%d %H:%M")])
        try: acad_list = json.loads(a.academic_json or "[]")
        except: acad_list = []
        if isinstance(acad_list, dict): acad_list = [acad_list]
        for r in acad_list:
            ws_acad.append([a.id, r.get("sr",""), r.get("qualification",""), r.get("college",""), r.get("year",""), r.get("grade","")])
        try: prof_list = json.loads(a.professional_json or "[]")
        except: prof_list = []
        if isinstance(prof_list, dict): prof_list = [prof_list]
        for r in prof_list:
            ws_prof.append([a.id, r.get("sr",""), r.get("company",""), r.get("designation",""), r.get("tenure",""), r.get("reason","")])
        try: fam_list = json.loads(a.family_json or "[]")
        except: fam_list = []
        if isinstance(fam_list, dict): fam_list = [fam_list]
        for r in fam_list:
            ws_fam.append([a.id, r.get("sr",""), r.get("name",""), r.get("relation",""), r.get("age",""), r.get("occupation","")])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    safe_status = f"_{status}" if status else ""
    fname = f"applications_{from_date_str}_to_{to_date_str}{safe_status}.xlsx"
    return send_file(out, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --------- Test email route ---------
@app.route('/test_email', methods=['GET','POST'])
def test_email():
    if request.method == 'POST':
        to_addr = request.form.get('to','').strip()
        if not to_addr:
            flash("Please provide an email", "warning")
            return redirect(url_for('test_email'))
        ok = send_email(to_addr, "Test — Accord Interview System", "<h3>Test email from your app</h3><p>If you received this, SMTP works.</p>")
        if ok:
            flash(f"Test email sent to {to_addr}", "success")
        else:
            flash("Test email failed — check logs and .env (app password).", "danger")
        return redirect(url_for('test_email'))
    return render_template('test_email.html')

# --------- Resume download ---------
@app.route('/uploads/<path:filename>')
def download_file(filename):
    return send_from_directory(UPLOAD_DIR, filename, as_attachment=True)

# --------- RUN APP ---------
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
